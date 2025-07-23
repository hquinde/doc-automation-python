import pandas as pd
from openpyxl import load_workbook

# -----------------------------
# Reusable Writing Function
# -----------------------------
def write_sample_data_by_id(ws, sample_names, data_rows, element_columns, header_row=1, data_start_col=2):
    sample_row_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_col=1):
        sample_id = row[0].value
        if sample_id:
            sample_row_map[str(sample_id).strip()] = row[0].row

    for sample_name, values in zip(sample_names, data_rows):
        target_row = sample_row_map.get(sample_name)
        if target_row:
            for j, val in enumerate(values):
                ws.cell(row=target_row, column=data_start_col + j, value=val)
        else:
            print(f"Sample '{sample_name}' not found in sheet — skipping.")

# -----------------------------
# Setup
# -----------------------------
file_path = "Raw Data.xlsx"
sheet_name = "Concentrations"
template_path = "Template Report.xlsx"
output_path = "Report_Output.xlsx"

std_names = [
    'Std1-5ppb', 'Std2-20ppb', 'Std3-50ppb',
    'Std4-200ppb', 'Std5-500ppb', 'Std6-2000ppb'
]

element_columns = [
    'Na 23\n(ug/L)', 'Mg 24\n(ug/L)', 'Al 27\n(ug/L)', 'Si 28\n(ug/L)',
    'P 31\n(ug/L)', 'K 39\n(ug/L)', 'Ca-43 43\nHelium KED\n(ug/L)',
    'Ca-43std 43\n(ug/L)', 'Ca-44 44\nHelium KED\n(ug/L)', 'Ca-44std 44\n(ug/L)',
    'Mn 55\n(ug/L)', 'Fe 57\n(ug/L)', 'Co 59\n(ug/L)', 'Ni 60\n(ug/L)',
    'Cu 63\n(ug/L)', 'Zn 68\n(ug/L)', 'Se 78\n(ug/L)', 'Se 82\n(ug/L)',
    'Sr 88\n(ug/L)', 'Mo 96\n(ug/L)', 'Cd 113\n(ug/L)',
    'Pb 206\n(ug/L)', 'Pb 207\n(ug/L)', 'Pb 208\n(ug/L)'
]

sample_map = {
    "QCS": "QCS 200PPB 1%NO3",
    "Ca Chk 500 ppb": "Ca check 500% NO3",
    "MDL": "MDL",
    "CCV1 200 ppb": "CCV1 200 ppb",
    "CCV2": "CCV2", "CCV3": "CCV3", "CCV4": "CCV4", "CCV5": "CCV5",
    "CCV6": "CCV6", "CCV7": "CCV7", "CCV8": "CCV8", "CCV9": "CCV9",
    "CCV10": "CCV10", "CCV11": "CCV11",
    "QCB": "QCB",
    "CCB1": "CCB1", "CCB2": "CCB2", "CCB3": "CCB3", "CCB4": "CCB4", "CCB5": "CCB5",
    "CCB6": "CCB6", "CCB7": "CCB7", "CCB8": "CCB8", "CCB9": "CCB9", "CCB10": "CCB10", "CCB11": "CCB11",
    "LCB1": "LCB1", "LCB2": "LCB2", "LCB3": "LCB3", "LCB4": "LCB4", "LCB5": "LCB5"
}

Raw_results_map = {
    "CCASE 1 (1.58, 8.22)": None,  # to be filled dynamically
    **{str(i): i for i in range(2, 50)}, 
    "50-1" : "50-1", 
    "50-2": "50-2",
    **{str(i): i for i in range(51, 104)}
}

# -----------------------------
# Read Data
# -----------------------------
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Dynamically fill CCASE 1 from Concentrations sheet
sample_ids = df["Sample Id"].dropna().astype(str).tolist()
for i, sid in enumerate(sample_ids):
    if sid.strip() == "2" and i > 0:
        Raw_results_map["CCASE 1 (1.58, 8.22)"] = sample_ids[i - 1].strip()
        break

# -----------------------------
# Calibration
# -----------------------------
calibration_df = df[df["Sample Id"].isin(std_names)].drop_duplicates("Sample Id")
calibration_data = calibration_df[element_columns].values.tolist()

# -----------------------------
# QAQC: BLK (Rinse after QCB)
# -----------------------------
blk_row = None
for i, sid in enumerate(sample_ids):
    if sid == "QCB":
        for j in range(i + 1, len(sample_ids)):
            if sample_ids[j] == "Rinse":
                blk_row = df.iloc[j]
                break
        break
blk_values = blk_row[element_columns].tolist() if blk_row is not None else [None] * len(element_columns)

# QAQC sample data
qaqc_data, sample_names = [], list(sample_map.keys())
for name in sample_names:
    match = df[df["Sample Id"] == sample_map[name]]
    qaqc_data.append(match[element_columns].iloc[0].tolist() if not match.empty else [None] * len(element_columns))
qaqc_data.insert(sample_names.index("QCB") + 1, blk_values)
sample_names.insert(sample_names.index("QCB") + 1, "Blk")

# -----------------------------
# Raw Results data
# -----------------------------
raw_sample_names = list(Raw_results_map.keys())
raw_data = []
for name in raw_sample_names:
    match = df[df["Sample Id"] == Raw_results_map[name]]
    raw_data.append(match[element_columns].iloc[0].tolist() if not match.empty else [None] * len(element_columns))

# -----------------------------
# Write to Excel
# -----------------------------
wb = load_workbook(template_path)
write_sample_data_by_id(wb["Calibration"], std_names, calibration_data, element_columns)
write_sample_data_by_id(wb["QAQC"], sample_names, qaqc_data, element_columns)
write_sample_data_by_id(wb["Raw Results"], raw_sample_names, raw_data, element_columns)
wb.save(output_path)
print(f"Data written to {output_path}")
